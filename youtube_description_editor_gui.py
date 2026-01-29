"""
YouTube Bulk Description Editor - GUI Version
==============================================
A graphical interface to bulk edit YouTube video descriptions.

Features:
- Search for videos matching patterns
- Select individual videos or all for replacement
- Progress bar for operations
- Rollback capability if errors occur during updates

Requirements:
1. Enable YouTube Data API v3 in Google Cloud Console
2. Create OAuth 2.0 credentials (Desktop App)
3. Download client_secret.json and place it in this directory
4. Install dependencies: pip install -r requirements.txt
"""

import os
import re
import pickle
import threading
import json
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

import FreeSimpleGUI as sg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Load environment variables
load_dotenv()

# YouTube API scopes
SCOPES = ["https://www.googleapis.com/auth/youtube.force-ssl"]

# Configuration
CLIENT_SECRETS_FILE = os.getenv("CLIENT_SECRETS_FILE", "client_secret.json")
TOKEN_FILE = "token.pickle"
BACKUP_FILE = "description_backups.json"


# =============================================================================
# YouTube API Functions
# =============================================================================

class YouTubeAPI:
    def __init__(self):
        self.youtube = None
        self.authenticated = False

    def authenticate(self):
        """Authenticate with YouTube API using OAuth 2.0."""
        credentials = None

        if os.path.exists(TOKEN_FILE):
            with open(TOKEN_FILE, "rb") as token:
                credentials = pickle.load(token)

        if not credentials or not credentials.valid:
            if credentials and credentials.expired and credentials.refresh_token:
                credentials.refresh(Request())
            else:
                if not os.path.exists(CLIENT_SECRETS_FILE):
                    raise FileNotFoundError(
                        f"{CLIENT_SECRETS_FILE} not found!\n\n"
                        "Please download OAuth credentials from Google Cloud Console."
                    )
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
                credentials = flow.run_local_server(port=0)

            with open(TOKEN_FILE, "wb") as token:
                pickle.dump(credentials, token)

        self.youtube = build("youtube", "v3", credentials=credentials)
        self.authenticated = True
        return True

    def get_all_videos(self, progress_callback=None):
        """Retrieve all videos from the channel."""
        if not self.authenticated:
            raise Exception("Not authenticated")

        videos = []
        next_page_token = None

        # Get uploads playlist ID
        channels_response = self.youtube.channels().list(
            part="contentDetails",
            mine=True
        ).execute()

        if not channels_response.get("items"):
            return []

        uploads_playlist_id = channels_response["items"][0]["contentDetails"]["relatedPlaylists"]["uploads"]

        while True:
            playlist_response = self.youtube.playlistItems().list(
                part="snippet",
                playlistId=uploads_playlist_id,
                maxResults=50,
                pageToken=next_page_token
            ).execute()

            for item in playlist_response.get("items", []):
                video_id = item["snippet"]["resourceId"]["videoId"]
                title = item["snippet"]["title"]
                videos.append({"id": video_id, "title": title})

            if progress_callback:
                progress_callback(len(videos))

            next_page_token = playlist_response.get("nextPageToken")
            if not next_page_token:
                break

        return videos

    def get_video_details(self, video_id):
        """Get full video details including description."""
        response = self.youtube.videos().list(
            part="snippet,status",
            id=video_id
        ).execute()

        if response.get("items"):
            return response["items"][0]
        return None

    def get_video_details_batch(self, video_ids):
        """Get video details for multiple videos in one API call (max 50)."""
        if not video_ids:
            return {}
        
        # YouTube API allows up to 50 IDs per request
        response = self.youtube.videos().list(
            part="snippet,status",
            id=",".join(video_ids[:50])
        ).execute()

        results = {}
        for item in response.get("items", []):
            results[item["id"]] = item
        return results

    def update_video_description(self, video_id, video_details, new_description):
        """Update a video's description."""
        snippet = video_details["snippet"]

        self.youtube.videos().update(
            part="snippet",
            body={
                "id": video_id,
                "snippet": {
                    "title": snippet["title"],
                    "description": new_description,
                    "tags": snippet.get("tags", []),
                    "categoryId": snippet["categoryId"]
                }
            }
        ).execute()
        return True


# =============================================================================
# Backup/Restore Functions
# =============================================================================

def load_backups():
    """Load existing backups from file."""
    if os.path.exists(BACKUP_FILE):
        with open(BACKUP_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_backup(video_id, title, old_description):
    """Save a backup of the original description."""
    backups = load_backups()
    backups[video_id] = {
        "title": title,
        "description": old_description,
        "backup_time": datetime.now().isoformat()
    }
    with open(BACKUP_FILE, "w", encoding="utf-8") as f:
        json.dump(backups, f, ensure_ascii=False, indent=2)


def restore_from_backup(youtube_api, video_id):
    """Restore a video's description from backup."""
    backups = load_backups()
    if video_id not in backups:
        return False, "No backup found for this video"

    video_details = youtube_api.get_video_details(video_id)
    if not video_details:
        return False, "Could not fetch video details"

    try:
        youtube_api.update_video_description(
            video_id, video_details, backups[video_id]["description"]
        )
        return True, "Restored successfully"
    except Exception as e:
        return False, str(e)


# =============================================================================
# Pattern Processing Functions
# =============================================================================

# URL status cache to avoid rechecking the same URLs
_url_cache = {}


def extract_urls(text):
    """Extract all URLs from text."""
    url_pattern = r'https?://[^\s<>"\')\]]+[^\s<>"\')\].,;:!?]'
    urls = re.findall(url_pattern, text)
    return urls


def check_url_status(url, timeout=15, use_cache=True):
    """
    Check if a URL is accessible.
    Returns: (status_code, error_message or None, from_cache)
    """
    # Check cache first
    if use_cache and url in _url_cache:
        status, error = _url_cache[url]
        return status, error, True
    
    try:
        # Use more complete browser headers to avoid being blocked
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
        }
        req = urllib.request.Request(url, headers=headers, method='HEAD')
        with urllib.request.urlopen(req, timeout=timeout) as response:
            result = (response.getcode(), None)
    except urllib.error.HTTPError as e:
        # Some sites block HEAD, try GET
        if e.code == 405:
            try:
                req = urllib.request.Request(url, headers=headers)
                with urllib.request.urlopen(req, timeout=timeout) as response:
                    result = (response.getcode(), None)
            except urllib.error.HTTPError as e2:
                result = (e2.code, str(e2.reason))
            except Exception as e2:
                result = (None, str(e2))
        else:
            result = (e.code, str(e.reason))
    except urllib.error.URLError as e:
        result = (None, str(e.reason))
    except Exception as e:
        result = (None, str(e))
    
    # Cache the result
    _url_cache[url] = result
    return result[0], result[1], False


def export_all_links_to_excel(videos_with_links, filename=None):
    """
    Export ALL links report to Excel file with status and video privacy.
    Returns the filename of the created file.
    """
    if filename is None:
        filename = f"links_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "All Links Report"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    broken_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Headers
    headers = ["Video Title", "Video URL", "Privacy", "Link in Description", "Status", "Status Code", "Error"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row = 2
    for video in videos_with_links:
        video_title = video["title"]
        video_url = f"https://www.youtube.com/watch?v={video['id']}"
        privacy = video.get("privacy", "Unknown")
        
        for url, status_code, error in video.get("all_links", []):
            ws.cell(row=row, column=1, value=video_title)
            ws.cell(row=row, column=2, value=video_url)
            ws.cell(row=row, column=3, value=privacy)
            ws.cell(row=row, column=4, value=url)
            
            # Determine status
            if status_code is not None and status_code < 400:
                status = "OK"
                fill = ok_fill
            else:
                status = "Broken"
                fill = broken_fill
            
            ws.cell(row=row, column=5, value=status)
            ws.cell(row=row, column=6, value=status_code if status_code else "Unreachable")
            ws.cell(row=row, column=7, value=error or "")
            
            # Apply fill to status column
            ws.cell(row=row, column=5).fill = fill
            
            row += 1
    
    # Auto-adjust column widths
    for col in range(1, 8):
        max_length = len(headers[col-1])
        for r in range(2, row):
            cell_value = ws.cell(row=r, column=col).value
            if cell_value:
                max_length = max(max_length, min(len(str(cell_value)), 60))
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2
    
    # Freeze header row and add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{row-1}"
    
    wb.save(filename)
    return filename


def check_video_needs_update(description, find_pattern):
    """
    Check if a video description contains the find pattern.
    Returns: (needs_update, list of issues found)
    """
    issues = []
    if find_pattern and find_pattern in description:
        issues.append("Contains pattern to replace")
    return len(issues) > 0, issues


def process_description(description, find_pattern, replace_with):
    """
    Process a video description - replace find_pattern with replace_with.
    Returns: (new_description, was_modified, modifications_made)
    """
    original = description
    modifications = []

    if find_pattern and find_pattern in description:
        description = description.replace(find_pattern, replace_with)
        modifications.append("Replaced pattern")

    return description, description != original, modifications


# =============================================================================
# GUI Application
# =============================================================================

class YouTubeDescriptionEditorGUI:
    def __init__(self):
        self.youtube_api = YouTubeAPI()
        self.videos = []
        self.videos_needing_update = []
        self.window = None
        self.find_pattern = ""
        self.replace_with = ""
        self._link_check_result = None  # For thread-safe popup handling

        # Set theme
        sg.theme("DarkBlue13")

    def create_main_window(self):
        """Create the main application window."""
        # Find/Replace input section
        find_replace_frame = [
            [sg.Text("Find (exact text to search for):", font=("Helvetica", 10, "bold"))],
            [sg.Multiline("", size=(60, 8), key="-FIND_PATTERN-", font=("Consolas", 9))],
            [sg.Text("Replace with:", font=("Helvetica", 10, "bold"))],
            [sg.Multiline("", size=(60, 8), key="-REPLACE_WITH-", font=("Consolas", 9))],
        ]

        # Column for video list with checkboxes
        video_list_column = [
            [sg.Text("Videos Matching Pattern:", font=("Helvetica", 12, "bold"))],
            [sg.Table(
                values=[],
                headings=["", "Video Title"],
                col_widths=[3, 60],
                auto_size_columns=False,
                justification="left",
                num_rows=12,
                key="-VIDEO_TABLE-",
                enable_events=True,
                enable_click_events=True,
                select_mode=sg.TABLE_SELECT_MODE_EXTENDED,
                font=("Helvetica", 10)
            )],
            [
                sg.Checkbox("Select All", key="-SELECT_ALL-", enable_events=True),
                sg.Text("", expand_x=True),
                sg.Text("Selected: 0", key="-SELECTED_COUNT-")
            ]
        ]

        # Preview column
        preview_column = [
            [sg.Text("Description Preview:", font=("Helvetica", 12, "bold"))],
            [sg.TabGroup([
                [
                    sg.Tab("Current", [[sg.Multiline("", size=(50, 12), key="-PREVIEW_CURRENT-", disabled=True, font=("Consolas", 9))]]),
                    sg.Tab("After Update", [[sg.Multiline("", size=(50, 12), key="-PREVIEW_NEW-", disabled=True, font=("Consolas", 9))]])
                ]
            ])]
        ]

        # Status and progress
        status_row = [
            [sg.Text("Status:", font=("Helvetica", 10, "bold")), 
             sg.Text("Not connected", key="-STATUS-", size=(70, 1))],
            [sg.ProgressBar(100, orientation="h", size=(80, 20), key="-PROGRESS-", visible=False)],
            [sg.Text("", key="-PROGRESS_TEXT-", visible=False)]
        ]

        # Buttons
        button_row = [
            [
                sg.Button("🔐 Connect", key="-CONNECT-", size=(12, 1)),
                sg.Button("🔍 Search", key="-SEARCH-", size=(10, 1), disabled=True),
                sg.Button("🔗 Check 404 Links", key="-CHECK_LINKS-", size=(14, 1), disabled=True),
                sg.Button("✏️ Update Selected", key="-UPDATE-", size=(15, 1), disabled=True),
                sg.Button("🔄 Restore Backup", key="-RESTORE-", size=(15, 1), disabled=True),
                sg.Button("❌ Exit", key="-EXIT-", size=(8, 1))
            ]
        ]

        # Left column with find/replace inputs
        left_column = [
            [sg.Frame("Find & Replace", find_replace_frame, font=("Helvetica", 11, "bold"))],
        ]

        # Right column with video list and preview
        right_column = [
            [sg.Column(video_list_column), sg.VerticalSeparator(), sg.Column(preview_column)]
        ]

        # Main layout - compact with no extra spacing
        layout = [
            [sg.Text("🎬 YouTube Bulk Description Editor", font=("Helvetica", 14, "bold")), sg.Push(), *button_row[0]],
            [sg.HorizontalSeparator()],
            [sg.Column(left_column, vertical_alignment="top"), sg.VerticalSeparator(), sg.Column(right_column, vertical_alignment="top")],
            [sg.HorizontalSeparator()],
            *status_row,
        ]

        return sg.Window(
            "YouTube Bulk Description Editor",
            layout,
            finalize=True,
            resizable=True,
            size=(1400, 600)
        )

    def update_status(self, message):
        """Update status text."""
        if self.window:
            self.window["-STATUS-"].update(message)
            self.window.refresh()

    def show_progress(self, visible=True):
        """Show or hide progress bar."""
        if self.window:
            self.window["-PROGRESS-"].update(visible=visible)
            self.window["-PROGRESS_TEXT-"].update(visible=visible)
            self.window.refresh()

    def update_progress(self, current, total, text=""):
        """Update progress bar."""
        if self.window:
            percentage = int((current / total) * 100) if total > 0 else 0
            self.window["-PROGRESS-"].update(percentage)
            self.window["-PROGRESS_TEXT-"].update(f"{text} ({current}/{total})")
            self.window.refresh()

    def connect_to_youtube(self):
        """Authenticate with YouTube."""
        self.update_status("Connecting to YouTube...")
        try:
            self.youtube_api.authenticate()
            self.update_status("✅ Connected to YouTube successfully!")
            self.window["-SEARCH-"].update(disabled=False)
            self.window["-CHECK_LINKS-"].update(disabled=False)
            self.window["-CONNECT-"].update(disabled=True)
            return True
        except FileNotFoundError as e:
            sg.popup_error(str(e), title="Authentication Error")
            self.update_status("❌ Connection failed - missing credentials")
            return False
        except Exception as e:
            sg.popup_error(f"Authentication failed: {e}", title="Error")
            self.update_status(f"❌ Connection failed: {e}")
            return False

    def search_videos(self):
        """Search for videos matching the find pattern."""
        # Get the find pattern from GUI
        self.find_pattern = self.window["-FIND_PATTERN-"].get()
        self.replace_with = self.window["-REPLACE_WITH-"].get()

        if not self.find_pattern.strip():
            sg.popup_error("Please enter a pattern to find.", title="Error")
            return

        self.update_status("Fetching videos from channel...")
        self.show_progress(True)

        try:
            # Get all videos (just IDs and titles - cheap API call)
            self.videos = self.youtube_api.get_all_videos(
                progress_callback=lambda count: self.update_progress(count, count, "Fetching video list")
            )

            if not self.videos:
                self.update_status("No videos found on channel")
                self.show_progress(False)
                return

            # Fetch video details in batches of 50 (saves API quota!)
            self.videos_needing_update = []
            total = len(self.videos)
            batch_size = 50

            for batch_start in range(0, total, batch_size):
                batch_end = min(batch_start + batch_size, total)
                batch_ids = [v["id"] for v in self.videos[batch_start:batch_end]]
                
                self.update_progress(batch_end, total, "Checking videos")

                # Fetch details for up to 50 videos in ONE API call
                details_batch = self.youtube_api.get_video_details_batch(batch_ids)

                for video in self.videos[batch_start:batch_end]:
                    video_details = details_batch.get(video["id"])
                    if not video_details:
                        continue

                    description = video_details["snippet"]["description"]
                    needs_update, issues = check_video_needs_update(description, self.find_pattern)

                    if needs_update:
                        self.videos_needing_update.append({
                            "id": video["id"],
                            "title": video["title"],
                            "issues": issues,
                            "details": video_details,
                            "selected": False
                        })

            # Update table
            self.update_video_table()

            self.show_progress(False)
            self.update_status(
                f"✅ Found {len(self.videos_needing_update)} videos matching pattern "
                f"(out of {len(self.videos)} total)"
            )

            if self.videos_needing_update:
                self.window["-UPDATE-"].update(disabled=False)
                self.window["-RESTORE-"].update(disabled=False)

        except Exception as e:
            self.show_progress(False)
            sg.popup_error(f"Error searching videos: {e}", title="Error")
            self.update_status(f"❌ Search failed: {e}")

    def update_video_table(self):
        """Update the video table with current data."""
        table_data = []
        for video in self.videos_needing_update:
            checkbox = "☑" if video["selected"] else "☐"
            title = video["title"][:70] + "..." if len(video["title"]) > 70 else video["title"]
            table_data.append([checkbox, title])

        self.window["-VIDEO_TABLE-"].update(values=table_data)
        self.update_selected_count()

    def update_selected_count(self):
        """Update the selected count display."""
        count = sum(1 for v in self.videos_needing_update if v["selected"])
        self.window["-SELECTED_COUNT-"].update(f"Selected: {count}")

    def toggle_video_selection(self, row_index):
        """Toggle selection for a video."""
        if 0 <= row_index < len(self.videos_needing_update):
            self.videos_needing_update[row_index]["selected"] = not self.videos_needing_update[row_index]["selected"]
            self.update_video_table()

    def select_all_videos(self, select):
        """Select or deselect all videos."""
        for video in self.videos_needing_update:
            video["selected"] = select
        self.update_video_table()

    def show_preview(self, row_index):
        """Show preview of current and new description."""
        if 0 <= row_index < len(self.videos_needing_update):
            video = self.videos_needing_update[row_index]
            current_desc = video["details"]["snippet"]["description"]
            
            # Check if this is a broken links result (has broken_links key)
            if "broken_links" in video:
                # Show broken links info in the "After Update" tab
                broken_info = "🔗 BROKEN LINKS FOUND:\n\n"
                for url, status, error in video["broken_links"]:
                    status_str = str(status) if status else "Unreachable"
                    broken_info += f"❌ [{status_str}] {url}\n   Error: {error}\n\n"
                self.window["-PREVIEW_CURRENT-"].update(current_desc)
                self.window["-PREVIEW_NEW-"].update(broken_info)
            else:
                new_desc, _, _ = process_description(current_desc, self.find_pattern, self.replace_with)
                self.window["-PREVIEW_CURRENT-"].update(current_desc)
                self.window["-PREVIEW_NEW-"].update(new_desc)

    def check_broken_links(self):
        """Check all videos for links and export full report."""
        self.update_status("Fetching videos from channel...")
        self.show_progress(True)

        try:
            # Get all videos
            self.videos = self.youtube_api.get_all_videos(
                progress_callback=lambda count: self.update_progress(count, count, "Fetching video list")
            )

            if not self.videos:
                self.update_status("No videos found on channel")
                self.show_progress(False)
                return

            # Fetch video details in batches and check links
            self.videos_needing_update = []
            all_videos_with_links = []  # For full Excel export
            total = len(self.videos)
            batch_size = 50
            total_links_checked = 0
            cached_links = 0
            broken_count = 0

            for batch_start in range(0, total, batch_size):
                batch_end = min(batch_start + batch_size, total)
                batch_ids = [v["id"] for v in self.videos[batch_start:batch_end]]
                
                self.update_progress(batch_end, total, "Fetching video details")

                details_batch = self.youtube_api.get_video_details_batch(batch_ids)

                for video in self.videos[batch_start:batch_end]:
                    video_details = details_batch.get(video["id"])
                    if not video_details:
                        continue

                    description = video_details["snippet"]["description"]
                    urls = extract_urls(description)
                    
                    # Get privacy status
                    privacy_status = video_details.get("status", {}).get("privacyStatus", "Unknown")
                    
                    if not urls:
                        continue

                    # Check each URL (with caching)
                    all_links = []
                    broken_links = []
                    for url in urls:
                        total_links_checked += 1
                        status_code, error, from_cache = check_url_status(url)
                        
                        if from_cache:
                            cached_links += 1
                            self.update_status(f"Link {total_links_checked} (cached): {url[:40]}...")
                        else:
                            self.update_status(f"Checking link {total_links_checked}: {url[:40]}...")
                        
                        # Store ALL links
                        all_links.append((url, status_code, error))
                        
                        if status_code is None or status_code >= 400:
                            broken_links.append((url, status_code, error))
                            broken_count += 1

                    # Add to full export list (all videos with links)
                    all_videos_with_links.append({
                        "id": video["id"],
                        "title": video["title"],
                        "privacy": privacy_status,
                        "all_links": all_links,
                        "broken_links": broken_links
                    })

                    # Add to GUI list only if has broken links
                    if broken_links:
                        self.videos_needing_update.append({
                            "id": video["id"],
                            "title": video["title"],
                            "issues": [f"{len(broken_links)} broken link(s)"],
                            "details": video_details,
                            "broken_links": broken_links,
                            "selected": False
                        })

            # Update table
            self.update_video_table()

            self.show_progress(False)
            
            # Always export ALL links to Excel (for filtering)
            if all_videos_with_links:
                excel_file = export_all_links_to_excel(all_videos_with_links)
                self.update_status(
                    f"✅ Checked {total_links_checked} links ({broken_count} broken, {cached_links} cached). Report: {excel_file}"
                )
                if self.videos_needing_update:
                    self.window["-RESTORE-"].update(disabled=False)
                # Store for main thread to show popup
                self._link_check_result = {
                    "videos": len(all_videos_with_links),
                    "broken": broken_count,
                    "links": total_links_checked,
                    "cached": cached_links,
                    "file": excel_file
                }
            else:
                self.update_status(
                    f"✅ No links found in {len(self.videos)} videos"
                )
                self._link_check_result = None

        except Exception as e:
            self.show_progress(False)
            self.update_status(f"❌ Link check failed: {e}")

    def update_selected_videos(self):
        """Update all selected videos."""
        selected = [v for v in self.videos_needing_update if v["selected"]]

        if not selected:
            sg.popup_error("No videos selected!", title="Error")
            return

        # Confirmation
        confirm = sg.popup_yes_no(
            f"Are you sure you want to update {len(selected)} video(s)?\n\n"
            "Backups will be created before updating.",
            title="Confirm Update"
        )

        if confirm != "Yes":
            return

        self.show_progress(True)
        self.update_status("Updating videos...")

        success_count = 0
        error_count = 0
        errors = []

        for i, video in enumerate(selected):
            self.update_progress(i + 1, len(selected), "Updating videos")

            try:
                # Get fresh video details
                video_details = self.youtube_api.get_video_details(video["id"])
                if not video_details:
                    errors.append(f"{video['title']}: Could not fetch details")
                    error_count += 1
                    continue

                current_desc = video_details["snippet"]["description"]

                # Create backup BEFORE updating
                save_backup(video["id"], video["title"], current_desc)

                # Process and update
                new_desc, was_modified, _ = process_description(current_desc, self.find_pattern, self.replace_with)

                if was_modified:
                    try:
                        self.youtube_api.update_video_description(
                            video["id"], video_details, new_desc
                        )
                        success_count += 1
                    except HttpError as e:
                        # Attempt to restore from backup
                        restore_success, restore_msg = restore_from_backup(
                            self.youtube_api, video["id"]
                        )
                        if restore_success:
                            errors.append(f"{video['title']}: Update failed, restored from backup")
                        else:
                            errors.append(f"{video['title']}: Update failed, restore also failed: {restore_msg}")
                        error_count += 1
                else:
                    success_count += 1  # Already up to date

            except Exception as e:
                errors.append(f"{video['title']}: {str(e)}")
                error_count += 1

        self.show_progress(False)

        # Show results
        result_msg = f"✅ Successfully updated: {success_count}\n❌ Errors: {error_count}"
        if errors:
            result_msg += "\n\nErrors:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                result_msg += f"\n... and {len(errors) - 10} more"

        sg.popup(result_msg, title="Update Complete")
        self.update_status(f"Update complete: {success_count} success, {error_count} errors")

        # Refresh the list
        self.search_videos()

    def restore_backup(self):
        """Show dialog to restore a video from backup."""
        backups = load_backups()

        if not backups:
            sg.popup("No backups available.", title="Restore Backup")
            return

        # Create list of backups
        backup_list = []
        for video_id, data in backups.items():
            title = data["title"][:50] + "..." if len(data["title"]) > 50 else data["title"]
            backup_time = data.get("backup_time", "Unknown")
            backup_list.append(f"{title} (ID: {video_id}) - Backed up: {backup_time}")

        # Show selection dialog
        layout = [
            [sg.Text("Select a backup to restore:")],
            [sg.Listbox(backup_list, size=(80, 15), key="-BACKUP_LIST-", select_mode=sg.LISTBOX_SELECT_MODE_SINGLE)],
            [sg.Button("Restore", key="-DO_RESTORE-"), sg.Button("Cancel")]
        ]

        restore_window = sg.Window("Restore Backup", layout, modal=True)

        while True:
            event, values = restore_window.read()

            if event in (sg.WIN_CLOSED, "Cancel"):
                break

            if event == "-DO_RESTORE-":
                selected = values["-BACKUP_LIST-"]
                if not selected:
                    sg.popup_error("Please select a backup to restore.")
                    continue

                # Extract video ID from selection
                selected_text = selected[0]
                video_id = selected_text.split("ID: ")[1].split(")")[0]

                confirm = sg.popup_yes_no(
                    f"Are you sure you want to restore the backup for this video?",
                    title="Confirm Restore"
                )

                if confirm == "Yes":
                    success, msg = restore_from_backup(self.youtube_api, video_id)
                    if success:
                        sg.popup("✅ Backup restored successfully!", title="Success")
                    else:
                        sg.popup_error(f"Failed to restore: {msg}", title="Error")
                break

        restore_window.close()

    def run(self):
        """Run the main application loop."""
        self.window = self.create_main_window()

        while True:
            event, values = self.window.read(timeout=100)

            if event in (sg.WIN_CLOSED, "-EXIT-"):
                break

            # Check for link check result from background thread
            if self._link_check_result is not None:
                result = self._link_check_result
                self._link_check_result = None
                sg.popup(
                    f"Link Check Complete!\n\n"
                    f"Videos with links: {result['videos']}\n"
                    f"Total links checked: {result['links']}\n"
                    f"Broken links: {result['broken']}\n"
                    f"Cached (skipped): {result['cached']}\n\n"
                    f"Excel report saved to:\n{result['file']}\n\n"
                    f"Use Excel filters to sort by Status, Privacy, etc.",
                    title="Links Report"
                )

            if event == "-CONNECT-":
                # Run authentication in main thread (needs browser)
                self.connect_to_youtube()

            elif event == "-SEARCH-":
                # Run search in a thread to keep UI responsive
                threading.Thread(target=self.search_videos, daemon=True).start()

            elif event == "-CHECK_LINKS-":
                # Run broken link check in a thread
                threading.Thread(target=self.check_broken_links, daemon=True).start()

            elif event == "-UPDATE-":
                self.update_selected_videos()

            elif event == "-RESTORE-":
                self.restore_backup()

            elif event == "-SELECT_ALL-":
                self.select_all_videos(values["-SELECT_ALL-"])

            elif isinstance(event, tuple) and event[0] == "-VIDEO_TABLE-":
                # Table click event
                if event[2][0] is not None:  # Row was clicked
                    row = event[2][0]
                    if event[2][1] == 0:  # Checkbox column clicked
                        self.toggle_video_selection(row)
                    self.show_preview(row)

            elif event == "-VIDEO_TABLE-":
                # Row selection changed
                if values["-VIDEO_TABLE-"]:
                    self.show_preview(values["-VIDEO_TABLE-"][0])

        self.window.close()


def main():
    app = YouTubeDescriptionEditorGUI()
    app.run()


if __name__ == "__main__":
    main()
